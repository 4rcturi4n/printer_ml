# src/printer_ml/dataset_maker.py
import os
import re
from dataclasses import dataclass
from typing import List, Optional, Tuple, Union
import numpy as np
import pandas as pd
from openpyxl import load_workbook


# -------------------- CONFIG TYPES --------------------

@dataclass(frozen=True)
class FolderPair:
    videos_folder: str
    xl_folder: str
    label: str


FolderPairLike = Union[FolderPair, Tuple[str, str, str]]


# ===================== PAIRING VIDEO AND XLS =====================

def p_from_mp4(path: str) -> Optional[float]:
    """
    Extract power from mp4 filename.
    Ignores big numbers (>=500) which are likely velocity (e.g. 1000).
    """
    base = os.path.basename(path)
    name_no_ext = os.path.splitext(base)[0]

    nums = re.findall(r"\d+[.,]?\d*", name_no_ext)
    for tok in nums:
        num_str = tok.replace(",", ".")
        try:
            val = float(num_str)
        except ValueError:
            continue

        if val < 500:
            return val

    print("WARNING: no valid power found in video name (only big numbers / velocity?):", base)
    return None


def p_from_xlsx(name: str) -> Optional[float]:
    """
    Extract power from excel filename.

    Works for:
      '11-15_1000.xlsx'      -> 15.0
      '13-14_1000.xlsx'      -> 14.0
      '16-12_5,1000.xlsx'    -> 12.5
      '21-10,1000.xlsx'      -> 10.0
    """
    base = os.path.basename(name)
    name_no_ext = os.path.splitext(base)[0]

    parts = name_no_ext.split("-")
    if len(parts) < 2:
        print("WARNING: unexpected excel name format:", base)
        return None

    second = parts[1]

    if "," in second:
        before_comma = second.split(",")[0]
        num_str = before_comma.replace("_", ".")
    else:
        before_underscore = second.split("_")[0]
        num_str = before_underscore

    try:
        return float(num_str)
    except ValueError:
        print("WARNING: bad power format in excel name:", base, "->", num_str)
        return None


def match_folder(
    videos: List[str],
    xl_names: List[str],
    xl_folder: str,
    label: str = "",
) -> Tuple[List[str], List[str], List[str]]:
    """
    Match video<->excel within SAME folder by power.

    - If a video has no excel with same power => dropped
    - If an excel has no video with same power => dropped
    - If multiple excels have same power => keep LAST one (by xl_names ordering)
    """
    excel_powers_by_name = {}
    for xn in xl_names:
        p = p_from_xlsx(xn)
        if p is None:
            print(f"[{label}] Delete EXCEL '{os.path.join(xl_folder, xn)}' because cannot parse power")
        else:
            excel_powers_by_name[xn] = p

    power_to_excel_names = {}
    for name, p in excel_powers_by_name.items():
        power_to_excel_names.setdefault(p, []).append(name)

    excel_by_power = {}
    for p, names in power_to_excel_names.items():
        names_sorted = sorted(names, key=lambda n: xl_names.index(n))
        if len(names_sorted) > 1:
            for del_name in names_sorted[:-1]:
                print(f"[{label}] Duplicate power={p}: delete first EXCEL '{os.path.join(xl_folder, del_name)}'")
        excel_by_power[p] = names_sorted[-1]

    video_powers = [p_from_mp4(v) for v in videos]

    matched_videos, matched_xl_names, matched_xl_paths = [], [], []
    used_powers = set()

    for v, pv in zip(videos, video_powers):
        if pv is None:
            print(f"[{label}] Delete VIDEO '{v}' because cannot parse power")
            continue

        xl_name = excel_by_power.get(pv)
        if xl_name is None:
            print(f"[{label}] Delete VIDEO '{v}' (power={pv}) because no Excel data in same folder")
        else:
            matched_videos.append(v)
            matched_xl_names.append(xl_name)
            matched_xl_paths.append(os.path.join(xl_folder, xl_name))
            used_powers.add(pv)

    for p, name in excel_by_power.items():
        if p not in used_powers:
            print(f"[{label}] Delete EXCEL '{os.path.join(xl_folder, name)}' (power={p}) because no video data in same folder")

    print(f"[{label}] Kept {len(matched_videos)} matched pairs.")
    return matched_videos, matched_xl_names, matched_xl_paths


# ===================== EXCEL READING =====================

def get_power_and_h25(xl_path: Optional[str]) -> Tuple[Optional[float], Optional[object]]:
    """
    Reads:
      - A1 for power (e.g. '15,5 mW')
      - H25 for axial resolution
    """
    if xl_path is None:
        return None, None

    wb = load_workbook(xl_path, data_only=True)
    ws = wb.active

    raw_a1 = str(ws["A1"].value)
    m = re.search(r"\d+[.,]?\d*", raw_a1)

    if m:
        num_str = m.group().replace(",", ".")
        try:
            power = float(num_str)
        except ValueError:
            power = None
    else:
        power = None

    axial = ws["H25"].value
    return power, axial


# ===================== MAIN BUILDER =====================

def build_dataset(
    folder_pairs: List[FolderPairLike],
    out_csv: str = "data/processed/video_xl_dataset.csv",
) -> pd.DataFrame:
    """
    Full pipeline:
      1) list videos/xlsx per folder pair
      2) match by power (same mechanics as your notebook)
      3) read A1 + H25
      4) build dataframe
      5) clean axial + log(axial)
      6) save CSV
    """

    video_list = []
    xlsx_aligned_names = []
    xl_aligned_paths = []

    for i, pair in enumerate(folder_pairs, start=1):
        if isinstance(pair, FolderPair):
            videos_folder, xl_folder, label = pair.videos_folder, pair.xl_folder, pair.label
        else:
            # must be a 3-tuple
            videos_folder, xl_folder, label = pair

        if not os.path.isdir(videos_folder):
            raise FileNotFoundError(f"Videos folder not found: {videos_folder}")
        if not os.path.isdir(xl_folder):
            raise FileNotFoundError(f"Excel folder not found: {xl_folder}")

        videos = [os.path.join(videos_folder, f) for f in os.listdir(videos_folder) if f.endswith(".mp4")]
        xls = [f for f in os.listdir(xl_folder) if f.endswith(".xlsx")]

        print(f"\nPAIR {i} [{label}] Found videos={len(videos)} excels={len(xls)}")

        v, xl_names, xl_paths = match_folder(videos, xls, xl_folder, label=label)

        video_list += v
        xlsx_aligned_names += xl_names
        xl_aligned_paths += xl_paths

    print("\nTOTAL matched videos:", len(video_list))
    print("TOTAL matched excels:", len(xl_aligned_paths))

    powers, axials = [], []
    for path in xl_aligned_paths:
        p, h = get_power_and_h25(path)
        powers.append(p)
        axials.append(h)

    df = pd.DataFrame({
        "video_name": [os.path.splitext(os.path.basename(v))[0] for v in video_list],
        "video_path": video_list,
        "xl_name": xlsx_aligned_names,
        "xl_path": xl_aligned_paths,
        "power_mW": powers,
        "axial_resolution": axials,
    })


    def make_video_id(path):
        folder = os.path.basename(os.path.dirname(path))  # 250617
        name   = os.path.splitext(os.path.basename(path))[0]  # p15.5_v1000_z4
        return f"{folder}\\{name}"

    df["video_id"] = df["video_path"].apply(make_video_id)

    # Clean + log
    df["axial_resolution"] = pd.to_numeric(df["axial_resolution"], errors="coerce")
    before = len(df)
    df = df[(df["axial_resolution"].notna()) & (df["axial_resolution"] != 0)].copy()
    df["log_axial_resolution"] = np.log(df["axial_resolution"])
    after = len(df)

    print(f"✅ Dropped {before - after} rows with invalid axial_resolution (#DIV/0! / None / 0). Kept {after} rows.")

    os.makedirs(os.path.dirname(out_csv), exist_ok=True)
    df.to_csv(out_csv, index=False)
    print("✅ Saved dataset to:", out_csv)

    return df
